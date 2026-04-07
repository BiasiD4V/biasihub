import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
wb = openpyxl.load_workbook(r'C:\Users\guilherme.moreira\Downloads\Banco_Atividade(2.0).xlsx', data_only=True)
ws = wb['Planilha1']

obras = []
current_obra = None
current_atv = None
state = 'idle'

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    vals = list(row) + [None]*6
    a, b, c, d, e, f = vals[0], vals[1], vals[2], vals[3], vals[4], vals[5]
    
    if isinstance(a, str) and a.strip().startswith('\u25b6'):
        current_obra = a.strip().replace('\u25b6', '').strip()
        state = 'obra'
        continue
    
    if a == 'ATIVIDADE' and c == 'JORNADA':
        state = 'expect_atv'
        continue
    
    if state == 'expect_atv' and isinstance(a, str) and a.startswith('['):
        jornada = str(c).replace('h','') if c else '8'
        current_atv = {
            'obra': current_obra or '',
            'atividade': a,
            'jornada': jornada,
            'unid': d or 'm',
            'qtd': e if e and e != '\u2014' else None,
            'tempoDias': f if f and f != '\u2014' else None,
            'profissionais': []
        }
        state = 'expect_prof_header'
        continue
    
    if state == 'expect_prof_header' and a == '#':
        state = 'reading_profs'
        continue
    
    if state == 'reading_profs':
        if a == 'TOTAL':
            if current_atv:
                current_atv['totalCoef'] = e if e and e != '\u2014' else None
                current_atv['totalHh'] = f if f and f != '\u2014' else None
                obras.append(current_atv)
            current_atv = None
            state = 'idle'
            continue
        if isinstance(a, (int, float)) and b:
            if current_atv:
                current_atv['profissionais'].append({
                    'profissao': b,
                    'unid': d or 'H',
                    'coef': e if e and e != '\u2014' else None,
                    'hhTotal': f if f and f != '\u2014' else None,
                })
            continue

from collections import OrderedDict
grouped = OrderedDict()
for atv in obras:
    ob = atv['obra']
    if ob not in grouped:
        grouped[ob] = []
    grouped[ob].append(atv)

for ob, atvs in grouped.items():
    print(f'\n--- {ob} ({len(atvs)} atividades) ---')
    for a in atvs:
        nome = a['atividade']
        nprof = len(a['profissionais'])
        thh = a['totalHh']
        print(f'  {nome} | {nprof} profs | TotalHh: {thh}')

print(f'\n\nTotal geral: {len(obras)} atividades')
print(json.dumps(obras[:2], ensure_ascii=False, indent=2))
