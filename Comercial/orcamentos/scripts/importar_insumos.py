"""
Importar BASE BIASI.xlsx para Supabase (tabelas insumos + insumos_historico)
Uso: python scripts/importar_insumos.py
"""
import os
import json
import hashlib
from datetime import datetime
import openpyxl

# ── Config Supabase ──
SUPABASE_URL = "https://vzaabtzcilyoknksvhrc.supabase.co"
SERVICE_KEY = os.environ.get(
    "SUPABASE_SERVICE_ROLE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZ6YWFidHpjaWx5b2tua3N2aHJjIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NDUyNDI0NiwiZXhwIjoyMDkwMTAwMjQ2fQ.b0QCcqqIJMrx8li0g_uRXoJ9z114YWyiHvu5QPjMG7o"
)

import urllib.request

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}


def supabase_post(table, rows):
    """Insert rows via REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    data = json.dumps(rows, default=str).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=HEADERS, method="POST")
    try:
        resp = urllib.request.urlopen(req)
        return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"  ❌ Erro HTTP {e.code}: {body[:200]}")
        return None


def supabase_get(table, params=""):
    """GET rows via REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{table}?{params}"
    req = urllib.request.Request(url, headers=HEADERS, method="GET")
    resp = urllib.request.urlopen(req)
    return json.loads(resp.read())


def gerar_codigo(fornecedor, descricao):
    """Gera código único baseado em fornecedor+descrição."""
    chave = f"{(fornecedor or '').strip().upper()}|{(descricao or '').strip().upper()}"
    return hashlib.md5(chave.encode()).hexdigest()[:12].upper()


def main():
    xlsx_path = r"C:\Users\Ryan\OneDrive - BIASI\BASE BIASI.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"❌ Arquivo não encontrado: {xlsx_path}")
        exit(1)

    print(f"📖 Lendo {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Planilha1"]

    # ── Ler todas as linhas ──
    rows_raw = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        fornecedor = (row[0] or "").strip() or None
        grupo = (row[1] or "").strip() if row[1] else None
        descricao = (row[2] or "").strip()
        unidade = (row[3] or "PÇ").strip()
        data_alteracao = row[4]
        custo = row[5]

        if not descricao:
            continue

        if isinstance(data_alteracao, datetime):
            data_str = data_alteracao.isoformat()
        elif data_alteracao:
            data_str = str(data_alteracao)
        else:
            data_str = None

        rows_raw.append({
            "fornecedor": fornecedor,
            "grupo": grupo,
            "descricao": descricao,
            "unidade": unidade,
            "data_alteracao": data_str,
            "custo": float(custo) if custo else 0,
        })

    print(f"📊 {len(rows_raw)} linhas lidas do Excel")

    # ── Deduplicar: mesmo fornecedor+descrição → pegar o mais recente ──
    insumos_map = {}
    historico_entries = []

    for r in rows_raw:
        codigo = gerar_codigo(r["fornecedor"], r["descricao"])

        if codigo not in insumos_map:
            insumos_map[codigo] = {
                "codigo": codigo,
                "descricao": r["descricao"],
                "unidade": r["unidade"],
                "fornecedor": r["fornecedor"],
                "grupo": r["grupo"],
                "custo_atual": r["custo"],
                "data_ultimo_preco": r["data_alteracao"],
            }
        else:
            existing = insumos_map[codigo]
            # Manter o mais recente como custo_atual
            if r["data_alteracao"] and (
                not existing["data_ultimo_preco"]
                or r["data_alteracao"] > existing["data_ultimo_preco"]
            ):
                existing["custo_atual"] = r["custo"]
                existing["data_ultimo_preco"] = r["data_alteracao"]

        # Cada entrada é um registro de histórico
        historico_entries.append({
            "codigo": codigo,
            "custo": r["custo"],
            "fornecedor": r["fornecedor"],
            "data_cotacao": r["data_alteracao"],
            "origem": "importacao_excel",
        })

    insumos_list = list(insumos_map.values())
    print(f"📦 {len(insumos_list)} insumos únicos ({len(historico_entries)} entradas de histórico)")

    # ── Inserir insumos em lotes de 200 ──
    print("\n🔄 Inserindo insumos no Supabase...")
    BATCH = 200
    inserted_count = 0
    failed_count = 0

    for i in range(0, len(insumos_list), BATCH):
        batch = insumos_list[i : i + BATCH]
        result = supabase_post("insumos", batch)
        if result:
            inserted_count += len(result)
        else:
            failed_count += len(batch)
        pct = min(100, round((i + len(batch)) / len(insumos_list) * 100))
        print(f"  [{pct:3d}%] {inserted_count} inseridos, {failed_count} falhas")

    print(f"✅ {inserted_count} insumos inseridos")

    # ── Buscar IDs dos insumos para vincular histórico ──
    print("\n🔄 Buscando IDs dos insumos...")
    # Buscar em lotes
    all_insumos = []
    offset = 0
    while True:
        batch = supabase_get("insumos", f"select=id,codigo&limit=1000&offset={offset}")
        if not batch:
            break
        all_insumos.extend(batch)
        offset += 1000
        if len(batch) < 1000:
            break

    codigo_to_id = {i["codigo"]: i["id"] for i in all_insumos}
    print(f"  {len(codigo_to_id)} insumos encontrados")

    # ── Inserir histórico ──
    print("\n🔄 Inserindo histórico de preços...")
    hist_rows = []
    for h in historico_entries:
        insumo_id = codigo_to_id.get(h["codigo"])
        if not insumo_id:
            continue
        hist_rows.append({
            "insumo_id": insumo_id,
            "custo": h["custo"],
            "fornecedor": h["fornecedor"],
            "data_cotacao": h["data_cotacao"],
            "origem": h["origem"],
        })

    hist_inserted = 0
    for i in range(0, len(hist_rows), BATCH):
        batch = hist_rows[i : i + BATCH]
        result = supabase_post("insumos_historico", batch)
        if result:
            hist_inserted += len(result)
        pct = min(100, round((i + len(batch)) / len(hist_rows) * 100))
        print(f"  [{pct:3d}%] {hist_inserted} registros de histórico")

    print(f"\n🎉 Importação concluída!")
    print(f"   {inserted_count} insumos")
    print(f"   {hist_inserted} registros de histórico")


if __name__ == "__main__":
    main()
